"""Turn the app's raw input into something an LLM can read (SPECS §7.2).

The app sends `input_type = text | image`. Binary documents (PDF, Excel) arrive as
`image` + base64. `/parse` also passes a `filename`, but `/enrich` does not — so we identify
binary documents by sniffing magic bytes, falling back to the filename extension. We route:

  - text                 -> ("text",   <string>)
  - PDF bytes            -> ("text",   <extracted text>)   via PyMuPDF
  - XLSX/XLS bytes       -> ("text",   <flattened rows>)   via openpyxl
  - real image bytes     -> ("vision", <data: URI>)        multimodal LLM

Heavy parsers are imported lazily so the module (and the test suite) load without them.
"""

from __future__ import annotations

import base64
import binascii
import logging

from ..schemas import InputType

logger = logging.getLogger("spendscope.preprocess")

Modality = str  # "text" | "vision"

_IMAGE_MIME = {
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "png": "image/png",
    "heic": "image/heic",
    "heif": "image/heif",
    "gif": "image/gif",
    "webp": "image/webp",
    "tiff": "image/tiff",
    "bmp": "image/bmp",
}


class PreprocessError(Exception):
    pass


def _ext(filename: str | None) -> str:
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


def _decode_base64(content: str) -> bytes:
    try:
        return base64.b64decode(content, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise PreprocessError(f"content is not valid base64: {exc}") from exc


def _detect_kind(data: bytes, ext: str) -> str:
    """Return 'pdf' | 'excel' | 'image' from magic bytes, with the extension as a tie-breaker."""
    if data[:4] == b"%PDF":
        return "pdf"
    # XLSX is a zip (PK\x03\x04); plain .xls is the old OLE compound format.
    if data[:4] == b"PK\x03\x04" and ext in {"", "xlsx", "xlsm", "xls", "zip"}:
        return "excel"
    if data[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":  # OLE2 (legacy .xls)
        return "excel"
    if ext == "pdf":
        return "pdf"
    if ext in {"xlsx", "xlsm", "xls"}:
        return "excel"
    return "image"


def _mime_for_image(data: bytes, ext: str) -> str:
    if data[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if data[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return "image/webp"
    return _IMAGE_MIME.get(ext, "image/jpeg")


def prepare(
    input_type: InputType, content: str, filename: str | None = None
) -> tuple[Modality, str]:
    """Return (modality, payload) ready to hand to the LLM router."""
    if input_type == InputType.text:
        return "text", content

    data = _decode_base64(content)
    ext = _ext(filename)
    kind = _detect_kind(data, ext)

    if kind == "pdf":
        return "text", _pdf_to_text(data)
    if kind == "excel":
        return "text", _excel_to_text(data)

    mime = _mime_for_image(data, ext)
    return "vision", f"data:{mime};base64,{content}"


def _pdf_to_text(data: bytes) -> str:
    try:
        import fitz  # PyMuPDF
    except ImportError as exc:  # pragma: no cover - depends on deployment
        raise PreprocessError("PyMuPDF not installed; cannot read PDF") from exc

    try:
        parts: list[str] = []
        with fitz.open(stream=data, filetype="pdf") as doc:
            for page in doc:
                parts.append(page.get_text("text"))
        text = "\n".join(parts).strip()
    except Exception as exc:  # noqa: BLE001
        raise PreprocessError(f"failed to read PDF: {exc}") from exc

    if not text:
        raise PreprocessError("PDF contained no extractable text (likely a scanned image)")
    return text


def _excel_to_text(data: bytes) -> str:
    try:
        import io

        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise PreprocessError("openpyxl not installed; cannot read spreadsheet") from exc

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(cells):
                    lines.append("\t".join(cells))
        wb.close()
        text = "\n".join(lines).strip()
    except Exception as exc:  # noqa: BLE001
        raise PreprocessError(f"failed to read spreadsheet: {exc}") from exc

    if not text:
        raise PreprocessError("spreadsheet contained no data")
    return text
